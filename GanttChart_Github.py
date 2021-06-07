import csv
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import argparse

#Define borders
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

#Gives default entry paramenters
def read_parameter():
    arguments = argparse.ArgumentParser(
        description="This generates a Gantt Chart given a personalised csv format")
    arguments.add_argument("-f", "--file", metavar="", type=str, dest= "filename", help="this is the input file", required=True)
    arguments.add_argument("-o", "--output", metavar="", type=str, dest="filename_output", help="this is the output file",
                           required=True)
    return arguments



#Open the input file for the Gantt Chart
def open_csv_as_dict(filename):
    dict = {}
    begin_date = datetime.datetime.strptime('01/01/2900', '%m/%d/%Y')
    finishing_date = datetime.datetime.strptime('01/01/1900', '%m/%d/%Y')
    with open(filename) as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            entry = {}
            entry['Description'] = row['Description']
            entry['Start.date'] = row['Start.date']
            entry['End.date'] = row['End.date']
            entry['Date_of_check_point'] = row['Date_of_check_point']
            entry['Check_point'] = row['Check_point']
            entry['Colour'] = row['Colour']
            entry['Note'] = row['Note']
            if not row['Task'] in dict:
                dict[row['Task']] = [entry]
            else:
                dict[row['Task']].append(entry)
            date_begin_row = datetime.datetime.strptime(row['Start.date'], '%m/%d/%Y')
            date_finishing_row = datetime.datetime.strptime(row['End.date'], '%m/%d/%Y')
            if date_begin_row < begin_date:
                begin_date = date_begin_row
            if date_finishing_row > finishing_date:
                finishing_date = date_finishing_row
    csvfile.close()
    return dict, begin_date, finishing_date

# Calculates total length of the project
def calculate_diff(dict, begin_date):
    month_max = 0
    for elem in dict:
        for item in dict[elem]:
            begin_date_row=datetime.datetime.strptime(item['Start.date'], '%m/%d/%Y')
            calculate_starting_month = ((begin_date_row.year - begin_date.year) * 12) + begin_date_row.month - begin_date.month+1
            item['month_start'] = calculate_starting_month
            end_date_row = datetime.datetime.strptime(item['End.date'], '%m/%d/%Y')
            calculate_finishing_month = ((end_date_row.year - begin_date.year) * 12) + end_date_row.month - begin_date.month + 1
            item['month_end'] = calculate_finishing_month
            if item['month_end'] > month_max:
                month_max = item['month_end']
            if item['Date_of_check_point'] != "No":
                checkpoint_date_row= datetime.datetime.strptime(item['Date_of_check_point'], '%m/%d/%Y')
                calculate_checkpoint_month = ((checkpoint_date_row.year - begin_date.year) * 12) + checkpoint_date_row.month - begin_date.month + 1
                item['month_checkpoint'] = calculate_checkpoint_month
    return dict, month_max

#Generates and complete the excel file that will be the Gantt Chart
def create_excel(dict, month_max, filename_output):
    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt_Chart"
    ws.cell(row=1, column=1, value="Tasks/Project Month")

    tot = month_max + 1
    for i in range(tot):
        ws.cell(row=1, column=i+1).border = thin_border
        if i != 0:
            ws.cell(row=1, column=i+1, value=i)

    line = 2
    for elem in dict:
        ws.cell(row = line, column=1, value=elem)
        ws.cell(row=line, column=1).border = thin_border
        for i in range(tot):
            ws.cell(row=line, column=i + 1).border = thin_border
            for entry in dict[elem]:
                if i >= entry['month_start'] and i <= entry['month_end']:
                    ws.cell(row=line, column=i + 1).fill = PatternFill(start_color=entry['Colour'], fill_type="solid")
                if 'month_checkpoint' in entry:
                    ws.cell(row=line, column=entry['month_checkpoint']+1, value=entry['Check_point'])

        line = line + 1
    wb.save(filename_output)

#Writes the completed input table
def write_output(dict, filename_output):
    output_file = open("GanttChat_input_complete.csv","w")
    output_file.write("Task,Description,Start.date,End.date,Month.start,Month.end,Date_of_check_point,Check_point,Month_checkpoint,Note\n")
    for elem in dict:
        for entry in dict[elem]:
            output_file.write(elem+",")
            output_file.write(entry['Description'] + ",")
            output_file.write(entry['Start.date']+",")
            output_file.write(entry['End.date'] +",")
            output_file.write(str(entry['month_start'])+",")
            output_file.write(str(entry['month_end'])+",")
            output_file.write(entry['Date_of_check_point']+",")
            output_file.write(entry['Check_point'] +",")
            if 'month_checkpoint' in entry:
                output_file.write(str(entry['month_checkpoint']) + ",")
            else:
                output_file.write("None,")
            output_file.write(str(entry['Note']) + ",")

            output_file.write("\n")


def main():
    args=read_parameter()
    arguments = args.parse_args()
    dict_input, begin_date, finishing_date = open_csv_as_dict(arguments.filename)
    dict_finale, month_max = calculate_diff(dict_input, begin_date)
    create_excel(dict_finale,month_max, arguments.filename_output)
    write_output(dict_finale, arguments.filename_output)

if __name__ == "__main__":
    main()